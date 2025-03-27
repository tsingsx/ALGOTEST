#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
文件作用：报告Agent模块，负责分析测试结果并生成测试报告
开发规划：实现基于大模型的测试结果分析和报告生成功能
"""

import os
import json
from typing import Dict, Any, List, TypedDict, Optional
from datetime import datetime
from loguru import logger
from langgraph.graph import StateGraph
from sqlalchemy.sql import text
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.config import get_settings, get_llm_config
from core.database import (
    get_db, 
    TestCase,
    update_test_case_status,
    update_test_task_status
)
from core.utils import generate_unique_id
from core.logger import get_logger
from core.llm import call_zhipu_api

# 获取带上下文的logger
log = get_logger("report_agent")

class ReportState(TypedDict):
    """报告Agent状态定义"""
    task_id: str  # 任务ID
    test_cases: Optional[List[Dict[str, Any]]]  # 测试用例列表
    analysis_results: Optional[List[Dict[str, Any]]]  # 分析结果列表
    report_data: Optional[Dict[str, Any]]  # Excel报告数据
    report_path: Optional[str]  # 报告文件路径
    errors: List[str]  # 错误信息
    status: str  # 任务状态


def analyze_test_results(state: ReportState) -> ReportState:
    """
    分析测试结果节点 - 从数据库读取测试用例结果并使用大模型分析
    
    Args:
        state: 当前状态
        
    Returns:
        更新后的状态
    """
    task_id = state['task_id']
    log.info(f"开始分析测试结果: {task_id}")
    
    try:
        # 从数据库获取该任务的所有测试用例
        with get_db() as db:
            cases = db.query(TestCase).filter(TestCase.task_id == task_id).all()
            if not cases:
                raise ValueError(f"未找到任务的测试用例: {task_id}")
            
            log.info(f"找到 {len(cases)} 个测试用例")
            
            # 获取LLM配置
            llm_config = get_llm_config()
            
            # 构建所有测试用例的信息
            test_cases_info = []
            for i, case in enumerate(cases, 1):
                input_data = case.input_data or {}
                expected_output = case.expected_output or {}
                actual_output = case.actual_output
                
                case_info = f"""
测试用例 {i}:
- 用例ID: {case.case_id}
- 名称: {input_data.get('name', '未命名')}
- 目的: {input_data.get('purpose', '无')}
- 测试步骤: {input_data.get('steps', '无')}
- 预期结果: {expected_output.get('expected_result', '无')}
- 验证方法: {expected_output.get('validation_method', '无')}
- 实际输出: {actual_output or '无输出'}
"""
                test_cases_info.append(case_info)
            
            # 构建整体提示词
            prompt = f"""
请分析以下所有测试用例的执行结果，判断每个测试用例是否通过，并提供详细的分析依据。

{os.linesep.join(test_cases_info)}

对每个测试用例，请提供以下内容：
1. 测试是否通过的判断（true/false）
2. 详细的分析依据，包括：
   - 对比预期结果和实际输出的差异
   - 根据验证方法进行的具体验证过程
   - 如果测试失败，指出具体的失败原因
3. 总结性结论

请按以下JSON格式输出，key为测试用例ID：
{{
    "test_case_1_id": {{
        "is_passed": true/false,
        "analysis": "详细的分析过程...",
        "conclusion": "总结性结论..."
    }},
    "test_case_2_id": {{
        "is_passed": true/false,
        "analysis": "详细的分析过程...",
        "conclusion": "总结性结论..."
    }},
    ...
}}
"""
            
            # 调用大模型API
            log.info("调用大模型API进行批量分析...")
            result = call_zhipu_api(prompt, llm_config)
            
            # 解析返回的JSON
            try:
                # 尝试直接解析
                analysis_results_data = json.loads(result)
            except json.JSONDecodeError:
                # 如果直接解析失败，尝试从文本中提取JSON
                import re
                json_match = re.search(r'\{[\s\S]*\}', result)
                if json_match:
                    analysis_results_data = json.loads(json_match.group())
                else:
                    raise ValueError("无法解析大模型返回的结果")
            
            # 存储分析结果
            analysis_results = []
            
            # 更新每个测试用例的结果
            for case in cases:
                case_analysis = analysis_results_data.get(case.case_id)
                if case_analysis:
                    # 更新测试用例的分析结果
                    case.result_analysis = case_analysis.get("analysis", "") + "\n\n" + case_analysis.get("conclusion", "")
                    case.is_passed = case_analysis.get("is_passed", False)
                    case.status = "completed"
                    
                    # 添加到分析结果列表
                    analysis_results.append({
                        "case_id": case.case_id,
                        "is_passed": case.is_passed,
                        "result_analysis": case.result_analysis
                    })
                    
                    log.info(f"测试用例 {case.case_id} 分析完成: {'通过' if case.is_passed else '未通过'}")
                else:
                    log.warning(f"未找到测试用例 {case.case_id} 的分析结果")
                    analysis_results.append({
                        "case_id": case.case_id,
                        "error": "未找到分析结果"
                    })
            
            # 提交所有更改
            db.commit()
            log.success(f"完成所有测试用例分析: {task_id}")
            
            # 更新状态
            return {
                **state,
                "test_cases": [
                    {
                        "case_id": case.case_id,
                        "input_data": case.input_data,
                        "expected_output": case.expected_output,
                        "actual_output": case.actual_output,
                        "is_passed": case.is_passed,
                        "result_analysis": case.result_analysis
                    }
                    for case in cases
                ],
                "analysis_results": analysis_results,
                "status": "analyzed"
            }
            
    except Exception as e:
        log.error(f"分析测试结果失败: {str(e)}")
        return {
            **state,
            "errors": state.get("errors", []) + [str(e)],
            "status": "error"
        }

def generate_excel_report(state: ReportState) -> ReportState:
    """
    生成Excel测试报告节点
    
    Args:
        state: 当前状态
        
    Returns:
        更新后的状态
    """
    task_id = state['task_id']
    log.info(f"开始生成Excel测试报告: {task_id}")
    
    try:
        # 确保report目录存在
        os.makedirs("data/report", exist_ok=True)
        
        # 从数据库获取所有测试用例和任务信息
        with get_db() as db:
            cases = db.query(TestCase).filter(TestCase.task_id == task_id).all()
            if not cases:
                raise ValueError(f"未找到任务的测试用例: {task_id}")
            
            # 获取任务信息
            task = db.execute(text("""
                SELECT algorithm_image, dataset_url 
                FROM test_tasks 
                WHERE task_id = :task_id
            """), {"task_id": task_id}).fetchone()
            
            if not task:
                raise ValueError(f"未找到任务信息: {task_id}")
            
            algorithm_image = task[0] or ""
            dataset_url = task[1] or ""
            
            # 获取LLM配置
            llm_config = get_llm_config()
            
            # 创建新的Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "测试报告"
            
            # 设置基本信息部分的样式
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置标题
            title = f"算法测试报告-{datetime.now().strftime('%Y_%m_%d')}"
            ws.merge_cells('A1:E1')
            title_cell = ws.cell(row=1, column=1, value=title)
            title_cell.font = Font(bold=True, size=12)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = thin_border
            
            # 添加基本信息
            info_data = [
                ["测试需求", "", "", "SDK版本版本", ""],
                ["中心授权版本", "", "", ""],
                ["测试人员", "", "", ""],
                ["EV_SDK镜像版本", algorithm_image, "", ""],
                ["数据集", dataset_url, "", ""],
                ["服务器配置", "", "", ""],
                ["数据集", "", "", ""],
                ["算法指标说明", "", "", ""]
            ]
            
            current_row = 2
            for info in info_data:
                # 如果是算法指标说明，合并整行
                if info[0] != "测试需求":
                    
                    ws.merge_cells(f'B{current_row}:E{current_row}')
                    cell = ws.cell(row=current_row, column=1, value=info[0])
                    cell.fill = header_fill
                    cell.border = thin_border
                    cell = ws.cell(row=current_row, column=2, value=info[1])
                    cell.border = thin_border
                else:
                    # 合并第3、4列
                    ws.merge_cells(f'B{current_row}:C{current_row}')
                    for col, value in enumerate(info, 1):  # 只取前3个值，因为第3、4列合并了
                        if col != 3:
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = thin_border
                            if col == 1:  # 第一列使用灰色背景
                                cell.fill = header_fill
                
                current_row += 1
            
            # 添加空行
            current_row += 1
            
            # 设置列宽
            ws.column_dimensions['A'].width = 20  # 分类
            ws.column_dimensions['B'].width = 25  # 子类
            ws.column_dimensions['C'].width = 40  # 标准
            ws.column_dimensions['D'].width = 15  # 测试结果
            ws.column_dimensions['E'].width = 50  # 备注
            
            # 添加精度测试结果标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value="精度测试结果")
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # 添加模型识别率测试分析标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value="模型识别率测试分析")
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # 添加性能测试分析标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value="性能测试分析")
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # 添加兼容性测试分析标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value="兼容性测试分析")
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # 添加规范测试分析标题
            ws.merge_cells(f'A{current_row}:E{current_row}')
            cell = ws.cell(row=current_row, column=1, value="规范测试分析")
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            current_row += 1
            
            # 设置规范测试表头
            headers = ["分类", "子类", "标准", "测试结果", "备注"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
            current_row += 1
            
            # 构建所有测试用例的信息
            test_cases_info = []
            for case in cases:
                input_data = case.input_data or {}
                name = input_data.get("name", "未命名测试用例")
                steps = input_data.get("steps", "")
                
                case_info = f"""
测试用例 {case.case_id}:
- 名称: {name}
- 步骤: {steps}
- 通过状态: {case.is_passed}
- 分析结果: {case.result_analysis or '无分析结果'}
"""
                test_cases_info.append(case_info)
            
            # 构建整体提示词
            prompt = f"""
请分析以下所有测试用例信息，为每个测试用例生成测试报告的一行数据。

{os.linesep.join(test_cases_info)}

对每个测试用例，请生成以下字段：
- category: 测试分类（如：功能测试、性能测试、接口测试等）
- sub_category: 具体测试的参数名称（从测试步骤中提取）
- standard: 该参数的作用和测试标准
- result: 根据is_passed确定（通过/不通过）
- note: 对result_analysis的简要总结

请按以下JSON格式返回，key为测试用例ID：
{{
    "test_case_id_1": {{
        "category": "分类名称",
        "sub_category": "参数名称",
        "standard": "参数作用和测试标准",
        "result": "通过/不通过",
        "note": "分析结果总结"
    }},
    "test_case_id_2": {{
        ...
    }}
}}
"""
            
            try:
                # 调用大模型API
                result = call_zhipu_api(prompt, llm_config)
                
                # 尝试解析JSON响应
                try:
                    # 直接尝试解析
                    report_data_all = json.loads(result)
                except json.JSONDecodeError:
                    # 如果直接解析失败，尝试从文本中提取JSON
                    import re
                    json_match = re.search(r'\{[\s\S]*\}', result)
                    if json_match:
                        report_data_all = json.loads(json_match.group())
                    else:
                        raise ValueError("无法从大模型响应中提取JSON数据")
                
                # 写入每个测试用例的数据
                for case in cases:
                    report_data = report_data_all.get(case.case_id)
                    if report_data:
                        # 写入Excel
                        ws.cell(row=current_row, column=1, value=report_data["category"])
                        ws.cell(row=current_row, column=2, value=report_data["sub_category"])
                        ws.cell(row=current_row, column=3, value=report_data["standard"])
                        ws.cell(row=current_row, column=4, value=report_data["result"])
                        ws.cell(row=current_row, column=5, value=report_data["note"])
                        
                        # 设置单元格样式
                        for col in range(1, 6):
                            cell = ws.cell(row=current_row, column=col)
                            cell.alignment = Alignment(wrap_text=True, vertical="center")
                            cell.border = thin_border
                            if col == 4:  # 测试结果列
                                if report_data["result"] == "通过":
                                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                else:
                                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        
                        current_row += 1
                    else:
                        log.warning(f"未找到测试用例 {case.case_id} 的报告数据")
                
            except Exception as e:
                log.error(f"处理测试报告数据时出错: {str(e)}")
                return {
                    **state,
                    "errors": state.get("errors", []) + [str(e)],
                    "status": "error"
                }
            
            # 保存Excel文件
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            report_path = f"data/report/test_report_{task_id}_{timestamp}.xlsx"
            wb.save(report_path)
            
            log.success(f"Excel报告生成成功: {report_path}")
            
            return {
                **state,
                "report_path": report_path,
                "status": "report_generated"
            }
            
    except Exception as e:
        log.error(f"生成Excel报告失败: {str(e)}")
        return {
            **state,
            "errors": state.get("errors", []) + [str(e)],
            "status": "error"
        }

def create_report_graph() -> StateGraph:
    """
    创建报告Agent工作流图
    
    Returns:
        工作流图
    """
    # 创建工作流图
    report_graph = StateGraph(ReportState)
    
    # 添加节点
    report_graph.add_node("analyze_test_results", analyze_test_results)
    report_graph.add_node("generate_excel_report", generate_excel_report)
    
    # 添加边
    report_graph.add_edge("analyze_test_results", "generate_excel_report")
    
    # 设置入口点和结束点
    report_graph.set_entry_point("analyze_test_results")
    report_graph.set_finish_point("generate_excel_report")
    
    return report_graph


async def run_report_generation(task_id: str) -> Dict[str, Any]:
    """
    运行报告生成Agent
    
    Args:
        task_id: 任务ID
        
    Returns:
        生成结果
    """
    # 创建工作流图
    report_graph = create_report_graph()
    
    # 编译工作流
    report_app = report_graph.compile()
    
    # 创建初始状态
    initial_state = {
        "task_id": task_id,
        "test_cases": None,
        "analysis_results": None,
        "errors": [],
        "status": "created"
    }
    
    # 运行工作流
    log.info(f"开始运行报告生成Agent: {task_id}")
    result = report_app.invoke(initial_state)
    log.info(f"报告生成Agent运行完成: {task_id}, 状态: {result['status']}")
    
    return result

if __name__ == "__main__":
    import asyncio
    import sys
    from core.logger import setup_logging
    
    # 设置日志
    setup_logging()
    
    # 使用指定的task_id
    task_id = "TASK1742525623_6049f9a3d00c"
    log.info(f"开始生成测试报告，任务ID: {task_id}")
    
    try:
        # 创建初始状态
        initial_state = {
            "task_id": task_id,
            "test_cases": None,
            "analysis_results": None,
            "errors": [],
            "status": "created"
        }
        
        # 直接调用generate_excel_report函数
        result = generate_excel_report(initial_state)
        
        # 检查结果
        if result["status"] == "report_generated":
            log.success(f"报告生成成功！")
            log.info(f"报告路径: {result.get('report_path')}")
        else:
            log.error(f"报告生成失败: {result.get('errors', ['未知错误'])}")
            sys.exit(1)
            
    except Exception as e:
        log.error(f"执行过程中出错: {str(e)}")
        sys.exit(1)

